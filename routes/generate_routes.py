from flask import Blueprint, request, render_template, send_file
from utils.scraper import setup_selenium_driver, scrape_report
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os
from utils.image_utils import download_image, create_image_zip, cleanup_directory
from urllib.parse import urlparse


generate_routes = Blueprint('generate_routes', __name__)

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def get_market_name_from_url(url):
    parsed_url = urlparse(url)
    return os.path.basename(parsed_url.path)

@generate_routes.route('/generate', methods=['POST'])
def generate_excel():
    urls = request.form.get('urls').strip().split('\n')
    urls = [url.strip() for url in urls if url.strip()]

    driver = setup_selenium_driver()

    scraped_data = []
    failed_urls = []
    image_urls = [] 
    for url in urls:
        try:
            report_data = scrape_report(url, driver)
            if report_data:
                scraped_data.append(report_data)
                if report_data.get("image_url"):
                    image_urls.append((report_data["image_url"], get_market_name_from_url(url)))
            else:
                failed_urls.append(url)
        except Exception as e:
            print(f"Error processing URL {url}: {str(e)}")
            failed_urls.append(url)
        
    driver.quit()
    cleanup_directory("images")
    downloaded_images = [download_image(img_url, name=market_name) for img_url, market_name in image_urls]
    image_zip_path = create_image_zip(downloaded_images)

    file_path = create_excel_report(scraped_data, failed_urls)
    return render_template('index.html', failed_urls=failed_urls, file_path=file_path,  image_zip=image_zip_path)


def create_excel_report(scraped_data, failed_urls):
    """Generates Excel report from the scraped data and saves it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    headers = ["Title", "Product Code", "URL", "Date", "Length", "Headline",
               "Price: Single User\nFormat: PDF & Excel", "Price: Site License\nFormat: PDF & Excel",
               "Price: Enterprise License\nFormat: PDF & Excel", "Description", "Table of Content",
               "Agenda / Schedule", "Executive Summary", "Sector", "Countries Covered",
               "Companies Mentioned", "Products Mentioned", "2022", "2023", "2031", "CAGR %", "Currency"]
    ws.append(headers)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = bold_font

    for row_idx, item in enumerate(scraped_data, start=2):
        ws.append([
            item.get("title", "N/A"),
            item.get("product_code", "N/A"),
            item.get("url", "N/A"),
            item.get("date", "N/A"),
            item.get("length", "N/A"),
            item.get("headline", "N/A"),
            item.get("price_single_user", "N/A"),
            item.get("price_site_license", "N/A"),
            item.get("price_enterprise_license", "N/A"),
            item.get("description", "N/A"),
            item.get("toc", "N/A"),
            item.get("agenda", "N/A"),
            item.get("executive_summary", "N/A"),
            item.get("sector", "N/A"),
            item.get("countries_covered", "N/A"),
            item.get("companies_mentioned", "N/A"),
            item.get("products_mentioned", "N/A"),
            item.get("data_2022", "N/A"),
            item.get("data_2023", "N/A"),
            item.get("data_2031", "N/A"),
            item.get("cagr", "N/A"),
            item.get("currency", "N/A")
        ])
        apply_error_formatting(ws, row_idx)

    if failed_urls:
        failed_ws = wb.create_sheet(title="Failed URLs")
        failed_ws.append(["Failed URLs"])
        for failed_url in failed_urls:
            failed_ws.append([failed_url])

    file_path = os.path.join(os.getcwd(), 'GII.xlsx')
    wb.save(file_path)
    return file_path

def apply_error_formatting(ws, row_idx):
    for cell in ws[row_idx]:
        if cell.value == "Error" or cell.value == "Report details not available.":
            for row_cell in ws[row_idx]:
                row_cell.fill = red_fill

                
@generate_routes.route('/download-images', methods=['GET'])
def download_images():
    zip_file_path = request.args.get('image_zip')
    return send_file(zip_file_path, as_attachment=True)

@generate_routes.route('/download')
def download_file():
    file_path = request.args.get('file_path')
    return send_file(file_path, as_attachment=True)


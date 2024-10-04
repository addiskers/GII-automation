from flask import Flask, render_template, send_file,request
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openai import OpenAI
from dotenv import load_dotenv
import os,re
from selenium import webdriver
from openpyxl import Workbook
from time import sleep
from openpyxl.styles import PatternFill, Font

load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")



def setup_selenium_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--headless')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-gpu')
    driver = webdriver.Chrome(options=options)
    
    return driver

app = Flask(__name__)

def AI(text, instruct):
    openai_client = OpenAI(api_key=api_key)

    prompt = f"{text}"
    instruction = f"""{instruct}"""

    completion = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": instruction},
            {"role": "user", "content": prompt},
        ],
    )
    assistant_response = completion.choices[0].message.content
    return assistant_response

def extract_bullet_points(ul_element, level=0):
    bullet_symbols = ["•", "o", ""]
    bullet_list = []
    for li in ul_element.find_all("li", recursive=False):
        bullet_symbol = bullet_symbols[min(level, len(bullet_symbols) - 1)]
        bullet_text = []
        for content in li.contents:
            if isinstance(content, str):
                text_content = content.strip()
                if text_content:
                    bullet_text.append(text_content)
            elif content.name == "strong":
                bullet_text.append(content.get_text(strip=True))
            elif content.name == "ul":
                if bullet_text:
                    bullet_list.append(
                        f"{bullet_symbol} {' '.join(bullet_text).strip()}"
                    )
                bullet_list.extend(extract_bullet_points(content, level + 1))
                bullet_text = []
        if bullet_text:
            bullet_list.append(f"{bullet_symbol} {' '.join(bullet_text).strip()}")
    return bullet_list






def extract_report_details(soup):
    try:
        description = soup.find("div", class_="report-details-description")
        first_para = description.find("p").text.strip()
        market_name = first_para.split("Market", 1)[0].strip()
        all_paragraphs = description.find_all("p")
        remaining_paragraphs = [para.text for para in all_paragraphs[1:]]
        remaining_text = "\n".join(remaining_paragraphs)
        remaining_text_instruction = "Rephrase as a market insights in 250 words in one paragraph"
        second_para = AI(remaining_text, remaining_text_instruction).strip()
        third_para = f"""
        Top-down and bottom-up approaches were used to estimate and validate the size of the {market_name} market and to estimate the size of various other dependent submarkets. The research methodology used to estimate the market size includes the following details: The key players in the market were identified through secondary research, and their market shares in the respective regions were determined through primary and secondary research. This entire procedure includes the study of the annual and financial reports of the top market players and extensive interviews for key insights from industry leaders such as CEOs, VPs, directors, and marketing executives. All percentage shares split, and breakdowns were determined using secondary sources and verified through Primary sources. All possible parameters that affect the markets covered in this research study have been accounted for, viewed in extensive detail, verified through primary research, and analyzed to get the final quantitative and qualitative data.
        """.strip()
        forth_para = f"{market_name} Market Segmental Analysis".strip()
        fifth_para = soup.select_one("#tab_default_1 > div:nth-of-type(3) > p").text.strip()
        sixth_para = f"Driver of the {market_name} Market".strip()
        ninth_div = soup.select_one("#tab_default_1 > div:nth-of-type(9)")
        driver_inst = f"Elaborate it as a market driver for {market_name} market in 100 words in one paragraph"
        if ninth_div:
            drivers = ninth_div.select_one("ul:nth-of-type(1)")
            if drivers:
                seventh_para = AI(drivers.text, driver_inst).strip()
        eighth_para = f"Restraints in the {market_name} Market".strip()
        ninth_inst = f"Elaborate it as a market restraint for {market_name} market in 100 words in one paragraph"
        if ninth_div:
            restraints_para = None
            for p in ninth_div.find_all("p"):
                if "restraint" in p.text.lower():
                    restraints_para = p
                    break

            if restraints_para:
                nextul = restraints_para.find_next_sibling("ul")
                ninth_para = AI(nextul.text, ninth_inst).strip()
        tenth_para = f"Market Trends of the {market_name} Market".strip()
        eleven_div = soup.select_one("#tab_default_1 > div:nth-of-type(11)")
        eleven_inst = f"Elaborate it as a market trend for {market_name} market in 100 words in one paragraph  "
        if eleven_div:
            ul_eleven = eleven_div.select_one("ul:nth-of-type(1)")
            if ul_eleven:
                ul_text = ul_eleven.text.strip()
                eleven_para = AI(ul_text, eleven_inst).strip()
            else:
                print("No <ul> found in the 11th <div>.")
        else:
            print("No 11th <div> found in #tab_default_1.")

        description_content = "\n\n".join(
            [
                first_para,
                second_para,
                third_para,
                forth_para,
                fifth_para,
                sixth_para,
                seventh_para,
                eighth_para,
                ninth_para,
                tenth_para,
                eleven_para,
            ]
        )  
        return description_content
    except Exception as e:
        print(f"Error extracting report details: {str(e)}")
        return "Report details not available."
    
# Function to scrape 
def scrape_report(url,driver):
    try:
        driver = driver
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "tabs-bar")))
        
        page_source1 = driver.page_source
        soup = BeautifulSoup(page_source1, "html.parser")
        toc_tab = driver.find_element(By.CSS_SELECTOR, "a[href='#tab_default_3']")
        driver.execute_script("arguments[0].scrollIntoView(true);", toc_tab)
        
        sleep(1)
        driver.execute_script("arguments[0].click();", toc_tab)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "tab_default_3")))
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, "special-toc-class")))
        page_source = driver.page_source
        soup_toc = BeautifulSoup(page_source, "html.parser")
        toc_section = soup_toc.find("div", {"class": "special-toc-class"})
        print(toc_section)
        if toc_section:
            ul_element = toc_section.find("ul")
            if ul_element:
                bullet_points = extract_bullet_points(ul_element)

        bullet_points_str = "\n".join(bullet_points)
        lines = bullet_points_str.strip().split("\n")
        toc_content = "\n".join(lines)
        if "• Introduction" not in toc_content or "o Objectives of the Study" not in toc_content:
            toc_content = "Error"
    except Exception as e:
            print(f"Error extracting table of contents for URL {url}: {str(e)}")
            
   
    summary = extract_report_details(soup)
    
    # Title
    head_div = soup.find("div", class_="d-sm-flex flex-sm-row-reverse align-items-center title")
    title = head_div.find("h1").text.strip() if head_div else ""
    
    # Product code and length
    code = soup.find("div", class_="report-segment-data max-width-640")
    report_id_tag = code.find("b", string="Report ID:") if code else None
    product_code = (
        report_id_tag.next_sibling.strip() if report_id_tag and report_id_tag.next_sibling else ""
    )
    product_code = re.sub(r"\W+", "", product_code)
    
    report_len_tag = code.find("b", string="Pages:") if code else None
    length = (
        report_len_tag.next_sibling.strip() if report_len_tag and report_len_tag.next_sibling else ""
    )
    length = re.sub(r"\D+", "", length)

    sector = soup.find("ol", class_="MuiBreadcrumbs-ol css-nhb8h9").find_all("li", class_="MuiBreadcrumbs-li")[1].text.strip() if soup.find("ol", class_="MuiBreadcrumbs-ol css-nhb8h9") else ""
    
    companies_list = []
    try:
        ten_div = soup.select_one("#tab_default_1 > div:nth-of-type(10)")
        companies_list = []
        if ten_div:
            ul_elements = ten_div.find_all("ul")
        for ul in ul_elements:
            preceding_p = ul.find_previous_sibling('p')
            if preceding_p and 'top player' in preceding_p.get_text(strip=True).lower():
                for li in ul.find_all("li"):
                    company = li.text.strip()
                    if company and company != '&nbsp;':
                        companies_list.append(f"◦ {company}")
            elif preceding_p and 'recent development' in preceding_p.get_text(strip=True).lower():
                break 
        cell_companies = "\n".join(companies_list)
    except Exception as e:
            print(f"Error extracting companies for URL {url}: {str(e)}")

        


    
    # Segments
    seg = soup.find("td", class_="fw-bold", string="Segments covered")
    segments_list = []
    try:
        if seg:
            next_td = seg.find_next_sibling()
            next_td_ul = next_td.find("ul")
            next_td_li = next_td_ul.find_all("li", recursive=False) if next_td_ul else []
            for li in next_td_li:
                main_category = li.contents[0].strip()
                subcategory = li.find("ul")
                if subcategory:
                    subcategory_items = [item.strip() for item in subcategory.stripped_strings]
                    subcategory_text = ", ".join(subcategory_items)
                    formatted_output = f"By {main_category} ({subcategory_text})"
                else:
                    formatted_output = f"By {main_category}"
                segments_list.append(formatted_output)
        products = ", ".join(segments_list)
    except Exception as e:
            print(f"Error extracting segments for URL {url}: {str(e)}")


    # Values and years
    first_para = soup.find("div", class_="report-details-description").find("p").text.strip()
    value_pattern = re.compile(r"USD (\d+\.?\d*)\s*(Billion|Million|Trillion|billion|million|trillion)")
    year_pattern = re.compile(r"\b(2022|2023|2031)\b")
    cagr_pattern = re.compile(r"CAGR of (\d+\.?\d*)\s*%")
    
    currency_values = value_pattern.findall(first_para)
    cagr = cagr_pattern.search(first_para)
    
    data_2022 = currency_values[0][0] if currency_values else None
    data_2023 = currency_values[1][0] if len(currency_values) > 1 else None
    data_2031 = currency_values[2][0] if len(currency_values) > 2 else None
    currency = "USD " + currency_values[0][1].title() if currency_values else None
    cagr_value = cagr.group(1) + "%" if cagr else ""
    
    # Countries covered
    countries_list = [
        "USA", "Canada", "Germany", "Spain", "Italy", "France", "UK", 
        "China", "India", "Japan", "South Korea", "Brazil", 
        "GCC Countries", "South Africa"
    ]
    formatted_countries = [f"◦ {country}" for country in countries_list]
    cell_countries = "\n".join(formatted_countries)
    
    # Prices
    price_single = "5300"
    price_sitelesense = "6200"
    price_enterprise = "7100"
    
    data.append([title, product_code, url, "", length, "", price_single, price_sitelesense, price_enterprise, summary, toc_content, "", "", sector, cell_countries, cell_companies, products, data_2022, data_2023, data_2031, cagr_value, currency])


# main page
@app.route('/')
def index():
    return render_template('index.html')

# generate Excel file
@app.route('/generate', methods=['POST'])
def generate_excel():
    driver = setup_selenium_driver()
    urls = request.form.get('urls').split('\n')
    urls = [url.strip() for url in urls if url.strip()]

    global data
    data = []  

    wb = Workbook()
    ws = wb.active
    headers = ["Title", "Product Code", "URL", "Date", "Length", "Headline", 
               "Price: Single User\nFormat: PDF & Excel", "Price: Site License\nFormat: PDF & Excel", 
               "Price: Enterprise License\nFormat: PDF & Excel", "Description", "Table of Content", 
               "Agenda / Schedule", "Executive Summary", "Sector", "Countries Covered", 
               "Companies Mentioned", "Products Mentioned", "2022", "2023", "2031", "CAGR %", "Currency"]
    ws.append(headers)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = bold_font

    failed_urls = []  

    for url in urls:
        try:
            scrape_report(url, driver) 
        except Exception as e:
            print(f"Error processing URL {url}: {str(e)}")
            failed_urls.append(url)  

    for row in data:
        ws.append(row)

    file_path = os.path.join(os.getcwd(), 'GII.xlsx')
    wb.save(file_path)

    return render_template('index.html', failed_urls=failed_urls, file_path=file_path)

@app.route('/download')
def download_file():
    file_path = request.args.get('file_path')
    return send_file(file_path, as_attachment=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
